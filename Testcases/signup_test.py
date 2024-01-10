import pytest
from Pages.FB_LoginPage import LoginPage
import openpyxl
import logging
from Utilities.logUtil import Logger

log = Logger(__name__,logging.INFO)

def get_data():
    workbook = openpyxl.load_workbook("C:\\Users\\pmrig\\OneDrive\\Desktop\\PDevelopment\\POM_design_framework\\Excel\\user_data.xlsx")
    sheet = workbook["LoginCred"]
    total_row = sheet.max_row
    total_column = sheet.max_column
    mainList = []
    for i in range(2, total_row+1):
        dataList = []
        for j in range(1, total_column+1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    return mainList 

@pytest.mark.usefixtures("get_browser")
@pytest.mark.parametrize("username,password" , get_data())
def test_dologin(username, password, get_browser):
    loginpage_object = LoginPage(get_browser)
    loginpage_object.fb_login(username, password)
    print(username, password)
    log.logger.info("Test Do Sign up successfully executed")